import os
import zipfile
from io import BytesIO

from django.contrib import admin
from django.http import HttpResponse
from django.utils import timezone
from django.utils.html import format_html
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from .models import Nazoratchi, Talabgor


@admin.register(Nazoratchi)
class NazoratchiAdmin(admin.ModelAdmin):
    list_display = ("id", "full_name", "username", "telegram_id", "status", "requested_at", "decided_at")
    list_filter = ("status",)
    search_fields = ("full_name", "username", "telegram_id")
    readonly_fields = ("requested_at",)


@admin.register(Talabgor)
class TalabgorAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "photo_thumbnail",
        "id_raqam",
        "familiya",
        "ism",
        "otasining_ismi",
        "telefon",
        "nazoratchi",
        "created_at",
    )
    search_fields = ("id_raqam", "familiya", "ism", "telefon")
    list_filter = ("nazoratchi",)
    readonly_fields = ("created_at", "photo_preview")
    fields = (
        "photo",
        "photo_preview",
        "familiya",
        "ism",
        "otasining_ismi",
        "telefon",
        "id_raqam",
        "nazoratchi",
        "created_at",
    )
    actions = ["export_to_excel", "export_photos_zip"]

    @admin.display(description="Rasm")
    def photo_thumbnail(self, obj: Talabgor) -> str:
        if not obj.photo:
            return "—"
        return format_html(
            '<img src="{}" style="height: 50px; width: 50px; object-fit: cover; '
            'border-radius: 4px;" />',
            obj.photo.url,
        )

    @admin.display(description="Rasm ko'rinishi")
    def photo_preview(self, obj: Talabgor) -> str:
        if not obj.photo:
            return "Rasm yuklanmagan"
        return format_html(
            '<img src="{}" style="max-height: 300px; max-width: 300px; '
            'border-radius: 6px;" />',
            obj.photo.url,
        )

    @staticmethod
    def _photo_filename(talabgor: Talabgor) -> str:
        """Rasm faylining ID raqamga moslashtirilgan nomi (masalan: 1234.jpg)."""
        _, ext = os.path.splitext(talabgor.photo.name)
        return f"{talabgor.id_raqam}{ext or '.jpg'}"

    @admin.action(description="Tanlangan talabgorlarni Excelga eksport qilish (rasm bilan)")
    def export_to_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Talabgorlar"

        headers = [
            "ID raqam",
            "Familiya",
            "Ism",
            "Otasining ismi",
            "Telefon",
            "Nazoratchi",
            "Qo'shilgan sana",
            "Rasm fayli nomi",
            "Rasm",
        ]
        ws.append(headers)

        widths = [12, 18, 15, 20, 15, 25, 20, 18, 16]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        row_num = 2
        for talabgor in queryset.order_by("id_raqam"):
            has_photo = bool(talabgor.photo) and os.path.exists(talabgor.photo.path)
            photo_filename = self._photo_filename(talabgor) if has_photo else ""

            ws.append(
                [
                    talabgor.id_raqam,
                    talabgor.familiya,
                    talabgor.ism,
                    talabgor.otasining_ismi,
                    talabgor.telefon,
                    str(talabgor.nazoratchi) if talabgor.nazoratchi else "",
                    timezone.localtime(talabgor.created_at).strftime("%Y-%m-%d %H:%M"),
                    photo_filename,
                    "",
                ]
            )
            ws.row_dimensions[row_num].height = 65

            if has_photo:
                try:
                    image = XLImage(talabgor.photo.path)
                    image.width = 80
                    image.height = 80
                    ws.add_image(image, f"I{row_num}")
                except Exception:
                    # Rasm formatini o'qib bo'lmasa, shu qatorni rasmsiz qoldiramiz.
                    pass

            row_num += 1

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"talabgorlar_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    @admin.action(description="Tanlangan talabgorlarning rasmlarini ZIP qilib yuklab olish (ID raqam nomi bilan)")
    def export_photos_zip(self, request, queryset):
        buffer = BytesIO()
        added = 0
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for talabgor in queryset.order_by("id_raqam"):
                if not talabgor.photo or not os.path.exists(talabgor.photo.path):
                    continue
                zip_file.write(talabgor.photo.path, arcname=self._photo_filename(talabgor))
                added += 1

        if added == 0:
            self.message_user(request, "Tanlangan talabgorlarda rasm topilmadi.")
            return None

        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type="application/zip")
        filename = f"talabgorlar_rasmlar_{timezone.now().strftime('%Y%m%d_%H%M%S')}.zip"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
